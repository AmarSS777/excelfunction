import openpyxl as xl

wb = xl.load_workbook("FDI.xlsx")
sheet = wb['Data']
tot = 0

for row in range(2, sheet.max_row):
    for column in range(2, sheet.max_column):
        cell = sheet.cell(row, column)
        tot = tot + cell.value
        average = tot/17
        average_cell = sheet.cell(row, 19)
        average_cell.value = average
    tot = 0
sheet['S1'] = "Average"

wb.save("FDI.xlsx")
